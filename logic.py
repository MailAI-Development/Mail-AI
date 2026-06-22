import requests
import openpyxl
import os
import re
import sys
import win32com.client
import pythoncom
from datetime import datetime
import pytz
import time
import csv
import json
import difflib
from math import log10, floor
import threading
import logging
import hmac as _hmac
import hashlib
import base64
try:
    import winreg  # Windows-only; used to persist trial start outside the app folder
except ImportError:
    winreg = None

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(os.path.join(os.path.expanduser("~"), "mailai.log"), encoding="utf-8"),
        logging.StreamHandler(),
    ]
)
logger = logging.getLogger(__name__)

API_KEY = "REPLACE_BEFORE_BUILD"

TRIAL_DAYS = 7
_LICENSE_SECRET = "REPLACE_BEFORE_BUILD"

APP_VERSION = "1.4"
GITHUB_REPO = "MailAI-Development/Mail-AI"
UPDATE_DOWNLOAD_URL = f"https://github.com/{GITHUB_REPO}/releases/latest/download/Mail.AI.{APP_VERSION}.exe"

class APIError(Exception):
    pass

def format_received_time(dt):
    s = dt.isoformat(timespec='minutes').replace('T', ' ')
    if len(s) > 16 and s[-6] in ('+', '-'):
        s = s[:-6] + ' ' + s[-6:]
    return s

def resource_path(relative_path):
    if getattr(sys, "frozen", False):
        base_path = getattr(sys, "_MEIPASS", os.path.dirname(sys.executable))
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

def data_path(relative_path):
    if getattr(sys, "frozen", False):
        base_path = os.path.dirname(sys.executable)
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

config_file = data_path("config.json")
duplicates_file = data_path("duplicates.json")
email_ids_file = data_path("email_ids.json")
custom_zones_file = data_path("custom_zones.json")
existing_vessels = {}
email_ids = set()
_email_ids_lock = threading.Lock()





TRANSLATIONS = {
    "English": {
        "welcome": "Welcome to Mail AI",
        "extract_something": "Extract something",
        "extract": "Extract",
        "listen": "Listen for emails",
        "filtering": "Filtering",
        "settings": "Settings",
        "filtering_settings": "Filtering settings",
        "email_caption": "Email address you want to extract data from:",
        "folder_caption": "Folder you want to extract data from:",
        "excel_caption": "Path to Excel spreadsheet (where data will be extracted to):",
        "clear_duplicates_caption": "Delete all existing duplicates stored",
        "clear_duplicates_btn": "Clear duplicates",
        "cleared": "✓ Duplicates cleared",
        "theme": "Theme:",
        "switch_light": "Switch to Light Mode",
        "switch_dark": "Switch to Dark Mode",
        "language": "Language:",
        "extract_page_header": "Extract something",
        "date_caption": "Choose a date to extract emails from. If left empty, all emails received today will be extracted:",
        "time_caption": "Choose a time to extract emails from. If left empty, all emails received between midnight and now will be extracted:",
        "start_extracting": "Start extracting",
        "tooltip": "Enter your email/folder/Excel path in filtering settings first",
        "no_email": "No email address has been defined, go to filtering settings",
        "email_extracting": "Email address currently extracting from:",
        "no_folder": "No folder has been defined, go to filtering settings",
        "folder_extracting": "Folder currently extracting from:",
        "no_excel": "No excel path has been defined, go to filtering settings",
        "excel_extracting": "Excel spreadsheet path:",
        "current_extraction": "Current extraction",
        "extraction_running": "Extraction is running",
        "extraction_stopped": "Extraction complete.",
        "continue_listen": "Continue listening",
        "donation_close": "Maybe later",
        "open_excel_btn": "Open spreadsheet",
        "excel_caption": "Path to Excel spreadsheet (leave empty to use default):",
        "extraction_complete_none": "Extraction complete. No results yielded.",
        "extraction_complete": "Extraction complete. Check Excel spreadsheet for results.",
        "stop_extracting": "Stop extracting",
        "new_extraction": "New extraction",
        "no_vessels": "No vessels found for the given date and time.",
        "vessels_extracted": "Vessels extracted:",
        "sender": "Sender",
        "subject": "Subject",
        "date": "Date",
        "location": "Location",
        "open_date": "Open Date",
        "build_year": "Built",
        "zone": "Zone",
        "listening_header": "Live listening",
        "listening_running": "Listening for emails...",
        "listening_paused": "Listening paused.",
        "pause_listen": "Pause listening",
        "resume_listen": "Resume listening",
        "listen_error": "Listening unable to run. Check filtering settings.",
        "outlook_not_running": "Check if Outlook is installed and running",
        "datetime_invalid": "Date/time is invalid! (check your format)",
        "excel_path_invalid": "Excel path is invalid! (check your format)",
        "folder_not_found": "Folder not found: ",
        "email_not_found": "Email address not found: ",
        "proxy_auth_error": "Service error. Please update the app or contact support.",
        "proxy_error_generic": "Extraction service error. Check your internet connection.",
        "setup_welcome_title": "Welcome to Mail AI",
        "setup_welcome_subtitle": "Let's get you set up in a few quick steps.",
        "setup_get_started": "Get started",
        "setup_email_title": "Your email address",
        "setup_email_desc": "Enter the Outlook email address you want to extract data from.",
        "setup_folder_title": "Outlook folder",
        "setup_folder_desc": "Enter the name of the Outlook folder to monitor for emails.",
        "setup_excel_title": "Excel spreadsheet",
        "setup_excel_desc": "Choose the Excel file where extracted data will be saved. Leave empty to use a default file in your Documents folder.",
        "setup_excel_browse": "Browse...",
        "setup_finish_title": "You're all set!",
        "setup_finish_desc": "Your configuration is saved. You can change these settings anytime from the Filtering page.",
        "setup_finish_btn": "Start using Mail AI",
        "setup_next": "Next",
        "setup_back": "Back",
        "setup_step": "Step",
        "custom_zones_header": "Custom Zone Mappings",
        "custom_zones_desc": "Add your own port-to-zone mappings:",
        "port_name_label": "Port Name:",
        "zone_label": "Zone:",
        "add_zone_btn": "Add Mapping",
        "zone_added": "Zone mapping added.",
        "zone_removed": "Zone mapping removed.",
        "zone_empty": "Please enter both a port name and zone.",
        "remove_zone_btn": "Remove",
        "no_custom_zones": "No custom zone mappings added yet.",
        "custom_zones_list": "Current custom mappings:",
        "limit_reached_title": "Free trial ended",
        "limit_reached_body": "Your 7-day free trial has ended. Upgrade to Pro for unlimited extractions — £9/month.",
        "upgrade_btn": "Upgrade to Pro — £9/month",
        "license_label": "License Key",
        "activate_btn": "Activate",
        "pro_active": "Pro active",
        "trial_active": "Free trial",
        "trial_days_left": "days left",
        "trial_expired": "Trial expired — upgrade to continue",
        "invalid_key": "Invalid or expired key",
        "pro_section_header": "Pro License",
    },
    "中文": {
        "welcome": "欢迎使用Mail AI",
        "extract_something": "提取数据",
        "extract": "提取",
        "listen": "监听邮件",
        "filtering": "筛选",
        "settings": "设置",
        "filtering_settings": "筛选设置",
        "email_caption": "您想提取数据的邮箱地址：",
        "folder_caption": "您想提取数据的文件夹：",
        "excel_caption": "Excel表格路径（数据将被提取到此处）：",
        "clear_duplicates_caption": "删除所有已存储的重复项",
        "clear_duplicates_btn": "清除重复项",
        "cleared": "✓ 重复项已清除",
        "theme": "主题：",
        "switch_light": "切换到浅色模式",
        "switch_dark": "切换到深色模式",
        "language": "语言：",
        "extract_page_header": "提取数据",
        "date_caption": "选择提取邮件的日期。如果留空，将提取今天收到的所有邮件：",
        "time_caption": "选择提取邮件的时间。如果留空，将提取从午夜到现在收到的所有邮件：",
        "start_extracting": "开始提取",
        "tooltip": "请先在筛选设置中填写邮箱/文件夹/Excel路径",
        "no_email": "未定义邮箱地址，请前往筛选设置",
        "email_extracting": "当前提取的邮箱地址：",
        "no_folder": "未定义文件夹，请前往筛选设置",
        "folder_extracting": "当前提取的文件夹：",
        "no_excel": "未定义Excel路径，请前往筛选设置",
        "excel_extracting": "Excel表格路径：",
        "current_extraction": "当前提取",
        "extraction_running": "提取进行中",
        "extraction_stopped": "提取完成。",
        "continue_listen": "继续监听",
        "donation_close": "也许以后",
        "open_excel_btn": "打开表格",
        "excel_caption": "Excel表格路径（留空则使用默认位置）：",
        "extraction_complete_none": "提取完成，未找到结果。",
        "extraction_complete": "提取完成，请查看Excel表格。",
        "stop_extracting": "停止提取",
        "new_extraction": "新建提取",
        "no_vessels": "在指定日期和时间内未找到船只。",
        "vessels_extracted": "已提取船只：",
        "sender": "发件人",
        "subject": "主题",
        "date": "日期",
        "location": "位置",
        "open_date": "开放日期",
        "build_year": "建造年份",
        "zone": "区域",
        "listening_header": "实时监听",
        "listening_running": "正在监听邮件...",
        "listening_paused": "监听已暂停。",
        "pause_listen": "暂停监听",
        "resume_listen": "恢复监听",
        "listen_error": "无法开始监听，请检查筛选设置。",
        "outlook_not_running": "请检查Outlook是否已安装并运行",
        "datetime_invalid": "日期/时间无效！（请检查格式）",
        "excel_path_invalid": "Excel路径无效！（请检查格式）",
        "folder_not_found": "未找到文件夹：",
        "email_not_found": "未找到邮箱地址：",
        "proxy_auth_error": "服务错误，请更新应用或联系支持。",
        "proxy_error_generic": "提取服务错误，请检查您的网络连接。",
        "setup_welcome_title": "欢迎使用 Mail AI",
        "setup_welcome_subtitle": "让我们通过几个简单的步骤完成设置。",
        "setup_get_started": "开始设置",
        "setup_email_title": "您的邮箱地址",
        "setup_email_desc": "输入您想要提取数据的 Outlook 邮箱地址。",
        "setup_folder_title": "Outlook 文件夹",
        "setup_folder_desc": "输入要监控的 Outlook 文件夹名称。",
        "setup_excel_title": "Excel 表格",
        "setup_excel_desc": "选择用于保存提取数据的 Excel 文件。留空将在文档文件夹中使用默认文件。",
        "setup_excel_browse": "浏览...",
        "setup_finish_title": "设置完成！",
        "setup_finish_desc": "您的配置已保存。您可以随时在筛选页面中更改这些设置。",
        "setup_finish_btn": "开始使用 Mail AI",
        "setup_next": "下一步",
        "setup_back": "上一步",
        "setup_step": "步骤",
        "custom_zones_header": "自定义区域映射",
        "custom_zones_desc": "添加您自己的港口-区域映射以补充内置的WPI数据库：",
        "port_name_label": "港口名称：",
        "zone_label": "区域：",
        "add_zone_btn": "添加映射",
        "zone_added": "区域映射已添加。",
        "zone_removed": "区域映射已删除。",
        "zone_empty": "请输入港口名称和区域。",
        "remove_zone_btn": "删除",
        "no_custom_zones": "尚未添加自定义区域映射。",
        "custom_zones_list": "当前自定义映射：",
        "limit_reached_title": "免费试用已结束",
        "limit_reached_body": "您的7天免费试用已结束。升级到专业版可无限提取 — £9/月。",
        "upgrade_btn": "升级到专业版 — £9/月",
        "license_label": "许可证密钥",
        "activate_btn": "激活",
        "pro_active": "专业版已激活",
        "trial_active": "免费试用",
        "trial_days_left": "天剩余",
        "trial_expired": "试用已过期 — 请升级以继续",
        "invalid_key": "无效或已过期的密钥",
        "pro_section_header": "专业版许可证",
    }
}

def t(key, language="English"):
    return TRANSLATIONS.get(language, TRANSLATIONS["English"]).get(key, key)


keywords = [
    'MV', 'DWT', 'dwt', 'open', 'vessel position', 
    'bulk carrier', 'handy', 'supramax', 'ultramax', 'panamax', 'kamsarmax', 'ETA', 'ETD'
]


_COUNTRY_ZONE = {
    'AU': 'AUS', 'NZ': 'AUS',
    'FI': 'BALTIC', 'SE': 'BALTIC', 'DK': 'BALTIC', 'PL': 'BALTIC',
    'EE': 'BALTIC', 'LV': 'BALTIC', 'LT': 'BALTIC',
    'UA': 'BSEA', 'RO': 'BSEA', 'BG': 'BSEA', 'GE': 'BSEA',
    'KZ': 'CIS', 'AZ': 'CIS', 'TM': 'CIS',
    'NL': 'CONTI', 'BE': 'CONTI', 'GB': 'CONTI', 'IE': 'CONTI',
    'NO': 'CONTI', 'IS': 'CONTI', 'PT': 'CONTI', 'DE': 'CONTI',
    'KE': 'EAFC', 'TZ': 'EAFC', 'MZ': 'EAFC', 'MG': 'EAFC',
    'ET': 'EAFC', 'SO': 'EAFC', 'KM': 'EAFC', 'SC': 'EAFC', 'MU': 'EAFC',
    'BR': 'ECSA', 'AR': 'ECSA', 'UY': 'ECSA',
    'CN': 'FE', 'JP': 'FE', 'KR': 'FE', 'TW': 'FE', 'KP': 'FE',
    'HK': 'FE', 'MO': 'FE',
    'IT': 'MED', 'GR': 'MED', 'HR': 'MED', 'SI': 'MED', 'ME': 'MED',
    'AL': 'MED', 'MT': 'MED', 'CY': 'MED', 'LB': 'MED', 'IL': 'MED',
    'LY': 'MED', 'TN': 'MED', 'DZ': 'MED', 'MA': 'MED', 'ES': 'MED',
    'TR': 'MED', 'EG': 'MED',
    'CO': 'NCSA', 'VE': 'NCSA', 'GY': 'NCSA', 'SR': 'NCSA',
    'AE': 'PG', 'KW': 'PG', 'IQ': 'PG', 'IR': 'PG', 'QA': 'PG',
    'BH': 'PG', 'OM': 'PG', 'SA': 'PG',
    'SD': 'RED SEA', 'YE': 'RED SEA', 'DJ': 'RED SEA', 'JO': 'RED SEA', 'ER': 'RED SEA',
    'ZA': 'SAFC', 'NA': 'SAFC',
    'SG': 'SEAS', 'MY': 'SEAS', 'ID': 'SEAS', 'PH': 'SEAS', 'VN': 'SEAS',
    'TH': 'SEAS', 'MM': 'SEAS', 'KH': 'SEAS', 'BN': 'SEAS', 'TL': 'SEAS',
    'CU': 'CARRIBEAN', 'JM': 'CARRIBEAN', 'TT': 'CARRIBEAN', 'BB': 'CARRIBEAN',
    'DO': 'CARRIBEAN', 'HT': 'CARRIBEAN', 'BS': 'CARRIBEAN', 'LC': 'CARRIBEAN',
    'VC': 'CARRIBEAN', 'GD': 'CARRIBEAN', 'AG': 'CARRIBEAN', 'DM': 'CARRIBEAN',
    'KN': 'CARRIBEAN', 'TC': 'CARRIBEAN',
    'NG': 'WAFC', 'GH': 'WAFC', 'CI': 'WAFC', 'CM': 'WAFC', 'AO': 'WAFC',
    'SN': 'WAFC', 'TG': 'WAFC', 'BJ': 'WAFC', 'GA': 'WAFC', 'GN': 'WAFC',
    'GW': 'WAFC', 'SL': 'WAFC', 'LR': 'WAFC', 'MR': 'WAFC', 'CV': 'WAFC',
    'GQ': 'WAFC', 'ST': 'WAFC', 'CD': 'WAFC', 'CG': 'WAFC',
    'GT': 'WCCA', 'SV': 'WCCA', 'HN': 'WCCA', 'NI': 'WCCA', 'CR': 'WCCA', 'PA': 'WCCA',
    'CL': 'WCSA', 'PE': 'WCSA',
    'PK': 'WCI', 'LK': 'ECI', 'BD': 'ECI', 'MV': 'WCI',
    'RU': 'CIS',
    'FR': 'MED',
    # --- additional coastal countries (single dominant zone) ---
    'GI': 'MED', 'MC': 'MED', 'BA': 'MED', 'SY': 'MED', 'PS': 'MED',
    'FO': 'CONTI', 'GL': 'CONTI', 'AX': 'BALTIC',
    'GM': 'WAFC', 'SH': 'WAFC',
    'RE': 'EAFC', 'YT': 'EAFC',
    'PG': 'SEAS',
    'FJ': 'AUS', 'NC': 'AUS', 'SB': 'AUS', 'VU': 'AUS', 'PF': 'AUS',
    'GU': 'FE',
    'BZ': 'CARRIBEAN', 'PR': 'CARRIBEAN', 'AW': 'CARRIBEAN', 'CW': 'CARRIBEAN',
    'KY': 'CARRIBEAN', 'VG': 'CARRIBEAN', 'VI': 'CARRIBEAN', 'AI': 'CARRIBEAN',
    'MS': 'CARRIBEAN', 'SX': 'CARRIBEAN', 'BQ': 'CARRIBEAN', 'BL': 'CARRIBEAN',
    'MF': 'CARRIBEAN', 'BM': 'CARRIBEAN',
    'EC': 'WCSA', 'FK': 'ECSA',
}

_US_STATE_ZONE = {
    'ME': 'USEC', 'NH': 'USEC', 'VT': 'USEC', 'MA': 'USEC', 'RI': 'USEC',
    'CT': 'USEC', 'NY': 'USEC', 'NJ': 'USEC', 'DE': 'USEC', 'MD': 'USEC',
    'VA': 'USEC', 'NC': 'USEC', 'SC': 'USEC', 'GA': 'USEC', 'PA': 'USEC',
    'FL': 'USG', 'AL': 'USG', 'MS': 'USG', 'LA': 'USG', 'TX': 'USG',
    'CA': 'USWC', 'OR': 'USWC', 'WA': 'USWC', 'AK': 'USWC', 'HI': 'USWC',
}

_IN_STATE_ZONE = {
    'GJ': 'WCI', 'MH': 'WCI', 'GA': 'WCI', 'KL': 'WCI', 'KA': 'WCI', 'DD': 'WCI',
    'WB': 'ECI', 'OD': 'ECI', 'OR': 'ECI', 'AP': 'ECI', 'TN': 'ECI', 'PY': 'ECI',
}

# Canada (2-letter province codes): Pacific vs Atlantic/Lakes.
_CA_STATE_ZONE = {
    'BC': 'USWC',
    'ON': 'USEC', 'QC': 'USEC', 'NL': 'USEC', 'NS': 'USEC',
    'NB': 'USEC', 'PE': 'USEC', 'MB': 'USEC',
}

# Mexico (3-letter state codes): Pacific vs Gulf vs Caribbean.
_MX_STATE_ZONE = {
    'BCN': 'WCCA', 'BCS': 'WCCA', 'SON': 'WCCA', 'SIN': 'WCCA', 'NAY': 'WCCA',
    'JAL': 'WCCA', 'COL': 'WCCA', 'MIC': 'WCCA', 'GRO': 'WCCA', 'OAX': 'WCCA',
    'CHP': 'WCCA',
    'TAM': 'USG', 'VER': 'USG', 'TAB': 'USG', 'CAM': 'USG', 'YUC': 'USG',
    'ROO': 'CARRIBEAN',
}

# Regional / broker shorthand → zone (unambiguous only; ambiguous terms left to log).
_REGION_ALIAS = {
    # --- broker shorthand / ranges ---
    'CJK': 'FE', 'FAR EAST': 'FE', 'F EAST': 'FE',
    'CJK-JP RANGE': 'FE', 'CJK-JAPAN RANGE': 'FE', 'CJK-JP': 'FE',
    'CJK-N.CHIAN RANGE': 'FE', 'CJK-N.CHINA RANGE': 'FE', 'NORTH OF CJK': 'FE',
    'N CHINA': 'FE', 'N.CHINA': 'FE', 'SOUTH CHINA': 'FE', 'S CHINA': 'FE',
    'HKG': 'FE', 'LYG': 'FE', 'LUOYUAN': 'FE',
    'SE ASIA': 'SEAS', 'SEASIA': 'SEAS', 'SOUTH EAST ASIA': 'SEAS', 'SPORE': 'SEAS',
    'SGP': 'SEAS', 'N.VIET': 'SEAS', 'N VIET': 'SEAS', 'E.MALAYSIA': 'SEAS',
    'E MALAYSIA': 'SEAS', 'ACEH, N. SUMATRA, INDO': 'SEAS', 'ACEH': 'SEAS',
    'ARA': 'CONTI', 'SKAW': 'CONTI',
    'GIB': 'MED', 'GIBRALTAR': 'MED', 'FULL MED': 'MED', 'MED RANGE': 'MED',
    'PMO': 'PG', 'PASSING MUSCAT': 'PG',
    'RIVER PLATE': 'ECSA', 'RIVERPLATE': 'ECSA',
    'RECALADA': 'ECSA', 'RECALADA (ARGENTINA)': 'ECSA',
    'ECI RANGE': 'ECI', 'WCI RANGE': 'WCI',
    'W. AFR ORDER': 'WAFC', 'W AFR': 'WAFC', 'W.AFR': 'WAFC', 'WEST AFRICA': 'WAFC',
    'AQABA': 'RED SEA', 'SIERRA LEONE': 'WAFC',
    # --- alternate / historical / transliterated port names ---
    'HOCHIMINH': 'SEAS', 'RANGOON': 'SEAS', 'MELAKA': 'SEAS',
    'KELANG': 'SEAS', 'PORT KELANG': 'SEAS', 'PORT KLANG': 'SEAS', 'KUANTAN': 'SEAS',
    'FUJARIAH': 'PG',
    'BAYRUT': 'MED', 'TARABULUS': 'MED', 'ANTWERP': 'CONTI',
    'CHITTAGONG': 'ECI', 'COCHIN': 'WCI', 'CALICUT': 'WCI', 'MORMUGAO': 'WCI',
    'VIZAG': 'ECI', 'PIPAVAV': 'WCI', 'PORT BLAIR': 'ECI',
}

def load_unlocode_dict():
    """Build port name → zone mapping from all three UN/LOCODE CSV parts."""
    mapping = {}
    folder = resource_path("csv")
    for i in range(1, 4):
        path = os.path.join(folder, f"UNLOCODE CodeListPart{i}.csv")
        if not os.path.exists(path):
            logger.warning(f"UN/LOCODE part {i} not found at {path}")
            continue
        with open(path, mode="r", encoding="latin-1") as f:
            for row in csv.reader(f):
                if len(row) < 7:
                    continue
                country = row[1].strip().upper()
                name = row[4].strip().upper()
                subdivision = row[5].strip().upper() if len(row) > 5 else ''
                func = row[6].strip()
                # Only seaports (function code starts with '1')
                if not func or func[0] != '1':
                    continue
                if not name or not country:
                    continue
                if country == 'US':
                    zone = _US_STATE_ZONE.get(subdivision)
                elif country == 'IN':
                    zone = _IN_STATE_ZONE.get(subdivision)
                elif country == 'CA':
                    zone = _CA_STATE_ZONE.get(subdivision)
                elif country == 'MX':
                    zone = _MX_STATE_ZONE.get(subdivision)
                else:
                    zone = _COUNTRY_ZONE.get(country)
                if zone and name not in mapping:
                    mapping[name] = [zone]
    logger.info(f"UN/LOCODE: loaded {len(mapping)} port entries")
    return mapping

def load_csv_into_dict(csv_file):
    mapping = {}
    with open(csv_file, mode="r", encoding="latin-1") as f:
        reader = csv.reader(f)
        next(reader, None)  # skip header if present
        for row in reader:
            if len(row) >= 2:  # avoid malformed lines
                key = row[0].strip().upper()
                value = row[1].strip()
                if key not in mapping:
                    mapping[key] = []
                if value not in mapping[key]:  # avoid duplicates
                    mapping[key].append(value)
    return mapping

def load_custom_zones():
    if os.path.exists(custom_zones_file):
        try:
            with open(custom_zones_file, "r", encoding="utf-8") as f:
                content = f.read().strip()
                if not content:
                    return {}
                data = json.loads(content)
                if not isinstance(data, dict):
                    return {}
                return data
        except (json.JSONDecodeError, OSError):
            return {}
    return {}

def save_custom_zones(zones):
    with open(custom_zones_file, "w", encoding="utf-8") as f:
        json.dump(zones, f, indent=1)

def add_custom_zone(port_name, zone):
    zones = load_custom_zones()
    key = port_name.strip().upper()
    zone = zone.strip().upper()
    if key not in zones:
        zones[key] = []
    if zone not in zones[key]:
        zones[key].append(zone)
    save_custom_zones(zones)

def remove_custom_zone(port_name):
    zones = load_custom_zones()
    key = port_name.strip().upper()
    if key in zones:
        del zones[key]
        save_custom_zones(zones)

def get_custom_zones_list():
    zones = load_custom_zones()
    return [(port, ", ".join(zone_list)) for port, zone_list in sorted(zones.items())]

def merge_custom_zones(csv_mapping):
    custom = load_custom_zones()
    for key, zone_list in custom.items():
        if key not in csv_mapping:
            csv_mapping[key] = []
        for zone in zone_list:
            if zone not in csv_mapping[key]:
                csv_mapping[key].append(zone)
    return csv_mapping

_PORT_NOISE = re.compile(
    r'^(PORT OF|PORT|ANCHORAGE(?: AT)?|TERMINAL|BERTH|ROADS?|OUTER|INNER|WEST|EAST|NORTH|SOUTH)\s+',
    re.IGNORECASE,
)

def _normalize_port(name):
    """Strip common prefixes/suffixes that obscure the core port name."""
    name = name.strip().upper()
    # Remove leading noise words iteratively (e.g. "PORT OF OUTER HAMBURG" → "HAMBURG")
    prev = None
    while prev != name:
        prev = name
        name = _PORT_NOISE.sub('', name).strip()
    return name

_ZONE_CODES = {
    'AUS', 'BALTIC', 'BSEA', 'CARRIBEAN', 'CIS', 'CONTI', 'EAFC',
    'ECI', 'ECSA', 'FE', 'MED', 'NCSA', 'PG', 'RED SEA', 'SAFC',
    'SEAS', 'USEC', 'USG', 'USWC', 'WAFC', 'WCCA', 'WCI', 'WCSA',
}

_COUNTRY_NAME_ZONE = {
    # Far East
    'CHINA': 'FE', 'JAPAN': 'FE', 'SOUTH KOREA': 'FE', 'KOREA': 'FE',
    'TAIWAN': 'FE', 'HONG KONG': 'FE', 'NORTH KOREA': 'FE',
    # South East Asia
    'SINGAPORE': 'SEAS', 'MALAYSIA': 'SEAS', 'INDONESIA': 'SEAS',
    'PHILIPPINES': 'SEAS', 'VIETNAM': 'SEAS', 'THAILAND': 'SEAS',
    'MYANMAR': 'SEAS', 'CAMBODIA': 'SEAS', 'BRUNEI': 'SEAS',
    # India (ambiguous — skip, brokers use WCI/ECI directly)
    # Persian Gulf
    'UAE': 'PG', 'UNITED ARAB EMIRATES': 'PG', 'SAUDI ARABIA': 'PG',
    'KUWAIT': 'PG', 'IRAQ': 'PG', 'IRAN': 'PG', 'QATAR': 'PG',
    'BAHRAIN': 'PG', 'OMAN': 'PG',
    # Mediterranean
    'ITALY': 'MED', 'GREECE': 'MED', 'SPAIN': 'MED', 'TURKEY': 'MED',
    'FRANCE': 'MED', 'CROATIA': 'MED', 'EGYPT': 'MED', 'LIBYA': 'MED',
    'TUNISIA': 'MED', 'ALGERIA': 'MED', 'MOROCCO': 'MED', 'ISRAEL': 'MED',
    'LEBANON': 'MED', 'CYPRUS': 'MED', 'MALTA': 'MED',
    # Continent
    'NETHERLANDS': 'CONTI', 'HOLLAND': 'CONTI', 'BELGIUM': 'CONTI',
    'GERMANY': 'CONTI', 'UK': 'CONTI', 'UNITED KINGDOM': 'CONTI',
    'NORWAY': 'CONTI', 'PORTUGAL': 'CONTI',
    # Baltic
    'FINLAND': 'BALTIC', 'SWEDEN': 'BALTIC', 'DENMARK': 'BALTIC',
    'POLAND': 'BALTIC', 'ESTONIA': 'BALTIC', 'LATVIA': 'BALTIC', 'LITHUANIA': 'BALTIC',
    # Black Sea
    'UKRAINE': 'BSEA', 'ROMANIA': 'BSEA', 'BULGARIA': 'BSEA', 'GEORGIA': 'BSEA',
    # CIS
    'RUSSIA': 'CIS', 'KAZAKHSTAN': 'CIS',
    # Red Sea
    'SUDAN': 'RED SEA', 'YEMEN': 'RED SEA', 'DJIBOUTI': 'RED SEA',
    'JORDAN': 'RED SEA', 'ERITREA': 'RED SEA',
    # East/South Africa
    'KENYA': 'EAFC', 'TANZANIA': 'EAFC', 'MOZAMBIQUE': 'EAFC',
    'MADAGASCAR': 'EAFC', 'SOMALIA': 'EAFC',
    'SOUTH AFRICA': 'SAFC', 'NAMIBIA': 'SAFC',
    # West Africa
    'NIGERIA': 'WAFC', 'GHANA': 'WAFC', 'IVORY COAST': 'WAFC',
    'COTE D\'IVOIRE': 'WAFC', 'CAMEROON': 'WAFC', 'ANGOLA': 'WAFC',
    'SENEGAL': 'WAFC', 'TOGO': 'WAFC', 'GABON': 'WAFC',
    # Australia
    'AUSTRALIA': 'AUS', 'NEW ZEALAND': 'AUS',
    # Americas
    'BRAZIL': 'ECSA', 'ARGENTINA': 'ECSA', 'URUGUAY': 'ECSA',
    'CHILE': 'WCSA', 'PERU': 'WCSA',
    'COLOMBIA': 'NCSA', 'VENEZUELA': 'NCSA',
    'MEXICO': 'WCCA', 'PANAMA': 'WCCA', 'GUATEMALA': 'WCCA',
    'CUBA': 'CARRIBEAN', 'JAMAICA': 'CARRIBEAN',
    # US (ambiguous — skip, brokers use USEC/USG/USWC directly)
}

_collapsed_index_cache = {}

def _collapse_name(s):
    """Strip everything but letters/digits, for spacing/punctuation-insensitive matching."""
    return re.sub(r'[^A-Z0-9]', '', s.upper())

def _get_collapsed_index(mapping):
    """Cached {collapsed_name -> set(zones)} view of the mapping (rebuilt if size changes)."""
    cached = _collapsed_index_cache.get(id(mapping))
    if cached is None or cached[0] != len(mapping):
        idx = {}
        for k, zones in mapping.items():
            ck = _collapse_name(k)
            if len(ck) >= 5:
                idx.setdefault(ck, set()).update(zones)
        _collapsed_index_cache[id(mapping)] = (len(mapping), idx)
        return idx
    return cached[1]

def lookup_value(input_text, mapping):
    input_text = input_text.strip().upper()

    if input_text in _ZONE_CODES:
        return input_text

    if input_text in _COUNTRY_NAME_ZONE:
        return _COUNTRY_NAME_ZONE[input_text]

    if input_text in _REGION_ALIAS:
        return _REGION_ALIAS[input_text]

    normalized = _normalize_port(input_text)
    candidates = [input_text] if input_text == normalized else [input_text, normalized]

    # Exact port match — return only if it resolves to a single unambiguous zone.
    for candidate in candidates:
        if candidate in mapping:
            zones = set(mapping[candidate])
            if len(zones) == 1:
                return next(iter(zones))
            # Multiple zones for one name (shared port name) — ambiguous, abstain.

    # Substring match — word-boundary only, keys >= 5 chars, single distinct zone only.
    for candidate in candidates:
        matched_zones = set()
        for key, zones in mapping.items():
            if len(key) < 5:
                continue
            key_re = r'\b' + re.escape(key) + r'\b'
            cand_re = r'\b' + re.escape(candidate) + r'\b'
            if re.search(key_re, candidate) or re.search(cand_re, key):
                matched_zones.update(zones)
        if len(matched_zones) == 1:
            return next(iter(matched_zones))
        # Zero matches → try next candidate; multiple distinct zones → ambiguous, abstain.

    # Space/punctuation-insensitive match — e.g. "HOCHIMINH" -> "HO CHI MINH CITY".
    collapsed_index = _get_collapsed_index(mapping)
    for candidate in candidates:
        cc = _collapse_name(candidate)
        if len(cc) < 5:
            continue
        exact = collapsed_index.get(cc)
        if exact and len(exact) == 1:
            return next(iter(exact))
        # Only candidate-is-substring-of-key (a shorter form of a fuller port name);
        # the reverse direction matches port names inside garbage strings, so it's excluded.
        matched = set()
        for ck, zones in collapsed_index.items():
            if cc in ck:
                matched.update(zones)
                if len(matched) > 1:
                    break
        if len(matched) == 1:
            return next(iter(matched))

    # Fuzzy match — accept only when unambiguous or a clear margin over a differing zone.
    for candidate in candidates:
        matches = difflib.get_close_matches(candidate, mapping.keys(), n=3, cutoff=0.88)
        if not matches:
            continue
        best = matches[0]
        best_zones = set(mapping[best])
        if len(best_zones) != 1:
            continue  # best match itself ambiguous
        best_zone = next(iter(best_zones))
        # Find the highest-ranked candidate that maps to a different zone.
        differing = next((m for m in matches if set(mapping[m]) != {best_zone}), None)
        if differing is None:
            return best_zone  # all close matches agree on the same zone
        best_ratio = difflib.SequenceMatcher(None, candidate, best).ratio()
        diff_ratio = difflib.SequenceMatcher(None, candidate, differing).ratio()
        if best_ratio - diff_ratio >= 0.05:
            return best_zone
        # Near-tie across different zones — ambiguous, abstain.

    try:
        with open(data_path("zone_misses.log"), "a", encoding="utf-8") as f:
            f.write(f"{input_text}\n")
    except OSError:
        pass

    return "UNKNOWN"

def is_relevant_email(email_subject, email_body):
    for keyword in keywords:
        if re.search(r'\b' + re.escape(keyword) + r'\b', email_subject, re.IGNORECASE):
            return True
    # only check body if subject didn't match, to avoid scanning long emails
    first_100_lines = '\n'.join(email_body.splitlines()[:100])
    for keyword in keywords:
        if re.search(r'\b' + re.escape(keyword) + r'\b', first_100_lines, re.IGNORECASE):
            return True
    return False

def get_first_n_lines(email_body):
    noise_patterns = re.compile(
        r'(?i)^('
        # quoted replies
        r'(>\s*)'
        # legal / email footer boilerplate
        r'|.*confidential.*'
        r'|.*disclaimer.*'
        r'|.*unsubscribe.*'
        r'|.*this (e-?mail|message) (is |was )?(intended|sent|confidential).*'
        r'|.*best regards.*|.*kind regards.*|.*warm regards.*'
        r'|.*sincerely.*|.*yours faithfully.*'
        r'|\s*sent from .*'
        r'|\s*get outlook for .*'
        # separator lines
        r'|-{3,}.*forwarded.*-{3,}'
        r'|-{5,}$'
        r'|_{5,}$'
        r'|={5,}$'
        r'|\*{5,}$'
        # bunker specs
        r'|.*\bLSFO\b.*\bISO\b.*'
        r'|.*\bMGO\b.*\bISO\b.*'
        r'|.*\bLSMGO\b.*\bISO\b.*'
        r'|.*bimco.*'
        r'|.*marpol.*'
        r'|.*sulphur content.*'
        r'|.*bunker.*quality.*'
        r'|.*bunker.*spec.*'
        r'|.*mixing of bunkers.*'
        r'|.*charterers.*warrant.*'
        r'|.*iso\s*8217.*'
        # speed / consumption tables
        r'|.*\bspeed\b.*\bcons\b.*'
        r'|.*abt\s+\d+.*knot.*'
        r'|.*\bbeaufort\b.*'
        r'|.*douglas sea.*'
        r'|.*good weather condition.*'
        r'|.*extrapolation.*'
        r'|.*positive current.*'
        r'|.*clean bottom.*'
        r'|.*adverse current.*'
        r'|.*wave.*swell.*'
        r'|.*ballast.*deballast.*'
        r'|.*lsmgo for exchanging.*'
        r'|.*liberty of using.*'
        r'|.*maneuvering.*'
        r'|.*maintenance works.*'
        # vessel particulars / tech specs
        r'|.*\bimo no\b.*'
        r'|.*\bcall sign\b.*'
        r'|.*port of registry.*'
        r'|.*\bflag\b.*\bclass\b.*'
        r'|.*\bbuilder\b.*'
        r'|.*date of delivery.*'
        r'|.*\bloa\b.*\blbp\b.*'
        r'|.*\bloa\b.*\bbeam\b.*'
        r'|.*\bgross tonnage\b.*'
        r'|.*\bgrt\b.*\bnrt\b.*'
        r'|.*\bgt\b.*\bnt\b.*\d+.*'
        r'|.*panama tonnage.*'
        r'|.*main engine.*'
        r'|.*auxiliary engine.*'
        r'|.*hatch cover.*'
        r'|.*tank top strength.*'
        r'|.*\bco2 devices\b.*'
        r'|.*radio remote control.*'
        r'|.*\bport idle\b.*'
        r'|.*\bport working\b.*'
        # last cargo / port history
        r'|.*last (five|ten|five|3|5|10)\s*(cargo|port).*'
        r'|.*last \d+\s*(cargo|port).*'
        r'|.*bunker on delivery.*'
        r'|.*buker on delivery.*'
        r'|all (above|details).*'
        # loadline / draft tables (never in summaries)
        r'|.*\bSSW\b.*'
        r'|.*\b(WINTER|SUMMER|TROPICAL)\b.*\bDWT\b.*'
        r'|.*\bDWT\b.*\bTPC\b.*'
        r'|.*\bTPC\b.*\bDRAFT\b.*'
        r'|.*\bDRAFT\b.*\bTPC\b.*'
        # fuel consumption lines
        r'|.*\bVLSFO\b.*\bLSMGO\b.*'
        r'|.*\bLSFO\b.*\bLSMGO\b.*'
        r'|.*\bMDO\b.*\bLSFO\b.*'
        # speed lines using K (knots abbrev) alongside a fuel type — e.g. "ABT 13K ON LSFO ABT 25MT"
        r'|.*\b\d+\s*k\b.*\b(lsfo|vlsfo|lsmgo|mgo)\b.*'
        # vessel registry / class
        r'|.*\bimo\b.*\b\d{7}\b.*'
        r'|.*\bimo number\b.*'
        r'|.*\bclass\b.*\b(BV|DNV|LR|ABS|NK|GL|CCS|RINA|KR)\b.*'
        r'|.*\bbuilt\b.*\bflag\b.*'
        # hold / capacity lines
        r'|.*\bgrain\b.*\bbale\b.*\bcbm\b.*'
        r'|.*\ball details about\b.*'
        r')'
    )

    relevant = re.compile(
        r'(?i)('
        r'\bM[./]?V[./]?\b|\bDWT\b|\b\d+\s*K\b|\bOPEN\b'
        r'|\bO\s*/\s*A\b|\bO\.A\.?\b|\bOA\b'
        r'|\b(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)\b'
        r'|\b(JANUARY|FEBRUARY|MARCH|APRIL|MAY|JUNE|JULY|AUGUST|SEPTEMBER|OCTOBER|NOVEMBER|DECEMBER)\b'
        r'|\d{1,2}[.)]\s'
        r')'
    )

    # Clean lines, tracking original indices
    cleaned = []
    for line in email_body.splitlines():
        stripped = line.strip()
        if not stripped or noise_patterns.match(stripped):
            cleaned.append(None)
        else:
            cleaned.append(stripped)

    # Mark relevant line indices, then expand by 1 context line each side
    relevant_indices = set()
    for i, line in enumerate(cleaned):
        if line and relevant.search(line):
            relevant_indices.add(i)

    context_indices = set()
    for i in relevant_indices:
        for j in (i - 1, i, i + 1):
            if 0 <= j < len(cleaned) and cleaned[j]:
                context_indices.add(j)

    result = [cleaned[i] for i in sorted(context_indices)]

    mv_name_re = re.compile(
        r'\bM\.?V\.?\b\s*["\']?\s*([A-Z][A-Z\s\d\-]{2,40}?)(?=["\']?\s*(?:\(|\bDWT\b|\bOPEN\b|\b\d{5}\b|,|$))',
        re.IGNORECASE
    )
    has_summary_context = re.compile(
        r'(?i)(\bOPEN\b|\bO\s*/\s*A\b|\bARD\b'
        r'|\b(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)\b'
        r'|\b(JANUARY|FEBRUARY|MARCH|APRIL|MAY|JUNE|JULY|AUGUST|SEPTEMBER|OCTOBER|NOVEMBER|DECEMBER)\b)'
    )
    seen_vessels = set()
    deduped = []
    for line in result:
        m = mv_name_re.search(line)
        if m:
            name_key = ' '.join(m.group(1).strip().upper().split())
            if name_key in seen_vessels and not has_summary_context.search(line):
                continue  # repeated vessel name with no new open/date info
            seen_vessels.add(name_key)
        deduped.append(line)
    result = deduped

    if not result:
        return None

    return '\n'.join(result)


def extract_details_from_email(preprocessed_body, csv_dict):
    payload = {
        "messages": [
            {"role": "system", "content":
            "You are a data extraction tool. Output ONLY the requested fields in the exact format shown."
            "No explanations, no notes, no extra text. If a value is unknown, write None."
            },
            {"role": "user", "content":
            f"Extract the following details for each vessel mentioned in the shipbroking email:\n\n"
            f"1. MV (Motor Vessel): vessel name, sometimes prefixed with MV but may be numbered with no MV, never include DWT — e.g. MV OCEAN STAR\n"
            f"2. Deadweight (DWT): DWT in K to 2 significant figures — e.g. 70K, 58K. None if unknown\n"
            f"3. Build Year: 4-digit year the vessel was built — often written alongside DWT as DWT/YEAR e.g. 57K/2012 but may be elsewhere. None if unknown\n"
            f"4. Vessel Open Location: the port or region where the vessel becomes available. Port name only in capitals, strip country and prefixes like OPEN/AT/IN/EX, strip suffixes like PORT. May be marked with O/A or phrased as 'Open <location>', 'EX SHIPYARD <location>', or 'delivery <location>'. If only a region code is given (e.g. CJK, ECI, WCI, NOPAC), return that. None if unknown\n"
            f"5. Vessel Open Date: the date the vessel becomes available. Date only, no year — e.g. 10 OCT, 20-22 NOV, EARLY OCT. May be phrased as open date, O/A, ARD, ETA, 'Expected time of delivery', or 'delivery date'. None if unknown\n"
            f"Format:\nMV: [MV name]\nDeadweight: [deadweight]\nBuild Year: [build year]\nVessel Open Location: [vessel open location]\nVessel Open Date: [vessel open date]\n"
            f"Repeat for each vessel mentioned in the email.\n"
            f"A vessel's details may span multiple lines.\n"
            f"Return None for any field that is missing.\n"
            f"Before moving on to the next vessel, separate each vessel with '---'.\n"
            f"Email:\n{preprocessed_body}\n\n"},
        ]
    }
    api_payload = {
        "model": "gpt-5.4-nano",
        "messages": payload["messages"],
    }
    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json",
    }
    last_error = None
    for attempt in range(3):
        try:
            resp = requests.post(
                "https://api.openai.com/v1/chat/completions",
                json=api_payload,
                headers=headers,
                timeout=(10, 45),  # (connect, read) — bounds a stalled call instead of hanging
            )
            resp.raise_for_status()
            data = resp.json()
            break
        except (requests.exceptions.RequestException, ValueError) as e:
            # ValueError covers json.JSONDecodeError (non-JSON body from a proxy/gateway error)
            last_error = e
            logger.warning(f"API request failed (attempt {attempt + 1}/3): {e}")
            if attempt < 2:
                time.sleep(2 ** attempt)
    else:
        raise APIError(str(last_error))

    try:
        details = data["choices"][0]["message"]["content"].strip()
    except (KeyError, IndexError, TypeError):
        logger.warning(f"Unexpected API response shape: {data}")
        raise APIError("malformed API response")

    vessels = details.split('---')
    extracted_vessels = []

    for vessel in vessels:
        mv = re.search(r'MV:\s*(.*)|MV\s*(.*)', vessel)
        deadweight = re.search(r'Deadweight:\s*(.*)', vessel)
        build_year = re.search(r'Build Year:\s*(.*)', vessel)
        location = re.search(r'Vessel Open Location:\s*(.*)', vessel)
        date_of_arrival = re.search(r'Vessel Open Date:\s*(.*)', vessel)

        def clean(value):
            if value is None:
                return None
            stripped = value.strip()
            return None if stripped.lower() == 'none' else stripped.upper()

        def clean_year(value):
            if value is None:
                return None
            stripped = value.strip()
            if stripped.lower() == 'none':
                return None
            match = re.match(r'(\d{4})', stripped)
            return match.group(1) if match else None

        vessel_data = {
            'MV': ensure_mv_prefix(clean(mv.group(1) if mv and mv.group(1) else (mv.group(2) if mv and mv.group(2) else None))),
            'Deadweight': normalize_dwt(clean(deadweight.group(1) if deadweight else None)),
            'Build Year': clean_year(build_year.group(1) if build_year else None),
            'Vessel Open Location': clean(location.group(1) if location else None),
            'Vessel Open Date': validate_date(clean(date_of_arrival.group(1) if date_of_arrival else None)),
        }

        open_location = vessel_data['Vessel Open Location']
        if open_location:
            zone = lookup_value(open_location, csv_dict)
            vessel_data['Zone'] = zone
        else:
            vessel_data['Zone'] = None

        extracted_vessels.append(vessel_data)
    return extracted_vessels



def resolve_excel_path(path):
    if not path or not path.strip():
        docs = os.path.join(os.path.expanduser("~"), "Documents")
        os.makedirs(docs, exist_ok=True)
        return os.path.join(docs, "Vessel_Data_Extraction.xlsx")
    path = os.path.normpath(path)
    if not path.lower().endswith(".xlsx"):
        path += ".xlsx"
    return path

def load_config():
    if os.path.exists(config_file):
        with open(config_file, "r", encoding="utf-8") as f:
            config = json.load(f)
        return config
    return {}

def save_config(config):
    with open(config_file, "w", encoding="utf-8") as f:
        json.dump(config, f, indent=1)

def load_email_ids():
    global email_ids
    with _email_ids_lock:
        if os.path.exists(email_ids_file):
            try:
                with open(email_ids_file, "r", encoding="utf-8") as f:
                    content = f.read().strip()
                    if not content:
                        return
                    email_ids = set(json.loads(content))
            except (json.JSONDecodeError, OSError):
                email_ids = set()


def save_email_ids():
    with _email_ids_lock:
        ids_list = list(email_ids)
        if len(ids_list) > 5000:
            ids_list = ids_list[-5000:]
            email_ids.clear()
            email_ids.update(ids_list)
    with open(email_ids_file, "w", encoding="utf-8") as f:
        json.dump(ids_list, f)



# Trial start is mirrored across three stores so deleting one (or reinstalling to a
# new folder) does not reset the trial. The effective start is the EARLIEST seen anywhere.
_TRIAL_REG_PATH = r"Software\MailAI"
_TRIAL_APPDATA_FILE = os.path.join(
    os.environ.get("LOCALAPPDATA") or os.path.expanduser("~"), "MailAI", "trial.dat"
)

def _read_reg_trial():
    if winreg is None:
        return None
    try:
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, _TRIAL_REG_PATH) as k:
            val, _ = winreg.QueryValueEx(k, "trial_start")
            return val
    except OSError:
        return None

def _write_reg_trial(date_str):
    if winreg is None:
        return
    try:
        with winreg.CreateKey(winreg.HKEY_CURRENT_USER, _TRIAL_REG_PATH) as k:
            winreg.SetValueEx(k, "trial_start", 0, winreg.REG_SZ, date_str)
    except OSError:
        pass

def _read_appdata_trial():
    try:
        with open(_TRIAL_APPDATA_FILE, "r", encoding="utf-8") as f:
            return f.read().strip()
    except OSError:
        return None

def _write_appdata_trial(date_str):
    try:
        os.makedirs(os.path.dirname(_TRIAL_APPDATA_FILE), exist_ok=True)
        with open(_TRIAL_APPDATA_FILE, "w", encoding="utf-8") as f:
            f.write(date_str)
    except OSError:
        pass

def _earliest_trial_start():
    """Earliest valid trial-start date found across config.json, registry and appdata."""
    starts = []
    for v in (load_config().get("trial_start"), _read_reg_trial(), _read_appdata_trial()):
        if v:
            try:
                starts.append(datetime.strptime(v, "%Y-%m-%d"))
            except ValueError:
                pass
    return min(starts) if starts else None


def refresh_access_state():
    """Record the trial start on first run (to all three stores), keep them in sync to the
    earliest known date, and revoke Pro if the stored key is invalid/expired. Idempotent and
    cheap — safe to call at startup and on each polling cycle."""
    config = load_config()
    changed = False

    earliest = _earliest_trial_start()
    earliest_str = (earliest or datetime.now()).strftime("%Y-%m-%d")
    # Propagate the earliest date to every store so removing one can't reset the trial.
    if config.get("trial_start") != earliest_str:
        config["trial_start"] = earliest_str
        changed = True
    if _read_reg_trial() != earliest_str:
        _write_reg_trial(earliest_str)
    if _read_appdata_trial() != earliest_str:
        _write_appdata_trial(earliest_str)

    stored_key = config.get("license_key", "")
    if config.get("is_pro", False) and not validate_license_key(stored_key):
        config["is_pro"] = False
        changed = True
    if changed:
        save_config(config)


def trial_days_left():
    earliest = _earliest_trial_start()
    if earliest is None:
        return TRIAL_DAYS
    elapsed = (datetime.now() - earliest).days
    return max(0, TRIAL_DAYS - elapsed)


def trial_active():
    return trial_days_left() > 0


def is_pro_active():
    return load_config().get("is_pro", False)


def access_allowed():
    """True if the user may extract — active Pro licence or an unexpired free trial."""
    return is_pro_active() or trial_active()


# ── Auto-update (via GitHub Releases) ───────────────────────────────
def _parse_version(v):
    v = (v or "").strip().lstrip("vV")
    parts = []
    for p in v.split("."):
        digits = "".join(ch for ch in p if ch.isdigit())
        parts.append(int(digits) if digits else 0)
    return tuple(parts) or (0,)


def check_for_update():
    """Return the latest version string if a newer release exists on GitHub, else None."""
    try:
        resp = requests.get(
            f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest",
            headers={"Accept": "application/vnd.github+json"},
            timeout=8,
        )
        resp.raise_for_status()
        tag = resp.json().get("tag_name", "")
        if tag and _parse_version(tag) > _parse_version(APP_VERSION):
            return tag.lstrip("vV")
    except Exception as e:
        logger.info(f"Update check skipped: {e}")
    return None


def cleanup_old_update():
    """Remove the leftover '.old' executable left by a previous self-update."""
    if not getattr(sys, "frozen", False):
        return
    old_path = sys.executable + ".old"
    if os.path.exists(old_path):
        try:
            os.remove(old_path)
        except OSError:
            pass


def _latest_exe_url():
    """Resolve the actual .exe asset URL from the latest GitHub release.

    Asset names are version-suffixed (e.g. Mail.AI.v1.4.exe), so a fixed
    /releases/latest/download/Mail.AI.exe URL would 404. Falls back to the
    hardcoded URL only if the API call fails to find an .exe asset."""
    try:
        resp = requests.get(
            f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest",
            headers={"Accept": "application/vnd.github+json"},
            timeout=15,
        )
        resp.raise_for_status()
        for asset in resp.json().get("assets", []):
            name = (asset.get("name") or "").lower()
            if name.endswith(".exe"):
                return asset["browser_download_url"]
    except Exception as e:
        logger.info(f"Could not resolve release asset via API: {e}")
    return UPDATE_DOWNLOAD_URL


def apply_update():
    """Download the latest exe, replace the running one, and relaunch. Frozen builds only.
    Raises on failure; on success it has already started the new exe (caller should quit)."""
    if not getattr(sys, "frozen", False):
        raise RuntimeError("Auto-update only applies to the built executable.")
    exe = sys.executable
    exe_dir = os.path.dirname(exe)
    new_path = os.path.join(exe_dir, "Mail.AI.new.exe")
    old_path = exe + ".old"

    download_url = _latest_exe_url()
    with requests.get(download_url, stream=True, timeout=180) as r:
        r.raise_for_status()
        with open(new_path, "wb") as f:
            for chunk in r.iter_content(chunk_size=1 << 16):
                if chunk:
                    f.write(chunk)

    if os.path.exists(old_path):
        try:
            os.remove(old_path)
        except OSError:
            pass
    os.rename(exe, old_path)      # Windows allows renaming a running exe
    os.rename(new_path, exe)      # move the freshly downloaded exe into place

    import subprocess
    subprocess.Popen([exe])      # launch the new version; caller then exits


def validate_license_key(key: str) -> bool:
    try:
        key = key.strip().upper()
        parts = key.split("-")
        if len(parts) != 3 or parts[0] != "MAILAI":
            return False
        year_month = parts[1]
        sig = parts[2]
        if len(year_month) != 6 or not year_month.isdigit():
            return False
        key_date = datetime.strptime(year_month, "%Y%m")
        if (datetime.now() - key_date).days > 35:
            return False
        expected = base64.b32encode(
            _hmac.new(_LICENSE_SECRET.encode(), year_month.encode(), hashlib.sha256).digest()
        )[:10].decode()
        return _hmac.compare_digest(sig, expected)
    except Exception:
        return False


def validate(date, time_str, email_address, folder, excel_path, outlook, language="English"):
    if outlook is None:
        return False, t("outlook_not_running", language), None
    try:
        if not date:
            date = str(datetime.now(pytz.UTC).date())
        if not time_str:
            time_str = "12:00 AM"
        sp_datetime = datetime.strptime(f"{date} {time_str}", '%Y-%m-%d %I:%M %p')

        specific_datetime = sp_datetime.replace(tzinfo=pytz.UTC)
        if specific_datetime > datetime.now(pytz.UTC):
            return False, t("datetime_invalid", language), None

    except ValueError:
        return False, t("datetime_invalid", language), None

    if excel_path and not excel_path.endswith(".xlsx"):
        return False, t("excel_path_invalid", language), None

    for account in outlook.Folders:
        if account.Name.upper() == email_address.upper():
            for subfolder in account.Folders:
                if subfolder.Name.upper() == folder.upper():
                    folder = subfolder
                    return True, "", specific_datetime
            else:
                return False, t("folder_not_found", language) + folder, None
    else:
        return False, t("email_not_found", language) + email_address, None
    

def is_excel_open(file_path):
    if not os.path.exists(file_path):
        return False
    try:
        with open(file_path, 'a+b'):
            return False
    except IOError:
        return True


def append_data_excel(file_path, data, specific_datetime, listening):
    headers = ['MV', 'DWT/Built', 'Vessel Open Location', 'Vessel Open Date', 'Zone', 'Sender', 'Subject', 'Received Time']

    today = datetime.now().date()
    sheet_name = today.strftime("%d %b %Y")

    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        sheet.append(headers)
        sheet.auto_filter.ref = "A1:H1"
        workbook.save(file_path)
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
    else:
        workbook = openpyxl.load_workbook(file_path)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(title=sheet_name)
            sheet.append(headers)
            sheet.auto_filter.ref = "A1:H1"

    for entry in data:
        subject = entry.get('Subject', '') or ''
        if len(subject) > 50:
            subject = subject[:50] + '...'
        dwt = entry.get('Deadweight', '') or ''
        year = entry.get('Build Year', '') or ''
        dwt_built = f"{dwt}/{year}" if dwt and year else (dwt or year or '')
        row = [
            entry.get('MV', ''),
            dwt_built,
            entry.get('Vessel Open Location', ''),
            entry.get('Vessel Open Date', ''),
            entry.get('Zone', ''),
            entry.get('Sender', ''),
            subject,
            entry.get('Received Time', ''),
        ]
        sheet.append(row)

    # Apply autofilter across the full used range (8 columns A–H).
    if sheet.max_row >= 1:
        sheet.auto_filter.ref = f"A1:H{sheet.max_row}"

    workbook.save(file_path)
    workbook.close()


def append_error_message(file_path, sender_email, email_subject):
    while is_excel_open(file_path):
        time.sleep(2)

    # Check if the Excel file exists
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
    else:
        # Open the existing workbook
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

    # Prepare the message to append
    error_message = f"Data not found: Email sender: {sender_email}, Email subject: {email_subject}"

    # Append the error message in a new row
    sheet.append([error_message])

    workbook.save(file_path)
    workbook.close()




def filter_data(vessels):
    return [v for v in vessels if v.get('MV') is not None]

def load_duplicates():
    if os.path.exists(duplicates_file):
        try:
            with open(duplicates_file, "r", encoding="utf-8") as f:
                content = f.read().strip()
                if not content:
                    return {}
                return json.loads(content)
        except json.JSONDecodeError:
            return {}
    return {}

def load_existing_vessels():
    global existing_vessels
    existing_vessels = load_duplicates()

def save_duplicates(duplicates):
    with open(duplicates_file, "w", encoding="utf-8") as f:
        json.dump(duplicates, f, indent=1)


def normalise_mv(name):
    if not name:
        return None
    name = re.sub(r'^(M\.?V\.?|M/V)\s*', '', str(name), flags=re.IGNORECASE).strip()
    name = re.sub(r'\s+\d+K.*$', '', name, flags=re.IGNORECASE).strip()
    return re.sub(r'\s+', ' ', name).upper()

def normalise_date(date):
    if not date:
        return None
    replacements = {
        'JANUARY': 'JAN', 'FEBRUARY': 'FEB', 'MARCH': 'MAR',
        'APRIL': 'APR', 'JUNE': 'JUN', 'JULY': 'JUL',
        'AUGUST': 'AUG', 'SEPTEMBER': 'SEP', 'OCTOBER': 'OCT',
        'NOVEMBER': 'NOV', 'DECEMBER': 'DEC'
    }
    for full, short in replacements.items():
        date = re.sub(rf'\b{full}\b', short, date, flags=re.IGNORECASE)
    date = re.sub(r'(\d+)(ST|ND|RD|TH)\b', r'\1', date, flags=re.IGNORECASE)
    date = re.sub(r'(\d+)\s+TO\s+(\d+)', r'\1-\2', date, flags=re.IGNORECASE)
    date = re.sub(r'\b(19|20)\d{2}\b', '', date)
    return date.strip().upper()

def ensure_mv_prefix(name):
    if not name:
        return None
    if not re.match(r'^MV\s+', name):
        return f"MV {name}"
    return name

def normalize_dwt(dwt_str):
    if not dwt_str:
        return None
    match = re.match(r'([\d,\.]+)\s*([KkMmTt]*)', dwt_str.strip())
    if not match:
        return None
    num = float(match.group(1).replace(',', ''))
    unit = match.group(2).upper()
    if 'MT' in unit or (num > 999 and 'K' in unit) or (num > 999 and not unit):
        num = num / 1000
    if num == 0:
        return None
    magnitude = floor(log10(abs(num)))
    rounded = round(num, -int(magnitude) + 1)
    return f"{int(rounded)}K"

def validate_date(date_str):
    if not date_str:
        return None
    date_str = normalise_date(date_str)
    months = r'(?:JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)'
    valid_patterns = [
        rf'^\d{{1,2}}\s+{months}$',                                                      # 10 OCT
        rf'^\d{{1,2}}-\d{{1,2}}\s+{months}$',                                            # 20-22 NOV
        rf'^\d{{1,2}}/\d{{1,2}}\s+{months}$',                                            # 10/11 OCT
        rf'^\d{{1,2}}\s+{months}\s*[-/]\s*\d{{1,2}}\s+{months}$',                       # 30 JUN - 2 JUL / 30 JUN / 2 JUL
        rf'^{months}\s+\d{{1,2}}$',                                                      # OCT 10
        rf'^(?:EARLY|MID|LATE|END)\s+{months}$',                                         # EARLY OCT
        rf'^(?:EARLY|MID|LATE|END)\s+{months}\s*/\s*(?:EARLY|MID|LATE|END)\s+{months}$', # LATE MAR / EARLY APR
    ]
    for pattern in valid_patterns:
        if re.match(pattern, date_str):
            return date_str
    return date_str

def is_valid_vessel(vessel):
    return (
        vessel.get('MV') not in (None, 'None') and
        vessel.get('Vessel Open Location') not in (None, 'None') and
        '->' not in str(vessel.get('Vessel Open Location', ''))
    )

def detect_duplicates(details):
    global existing_vessels
    vessels = []

    for vessel in details:
        if not is_valid_vessel(vessel):
            continue

        key = normalise_mv(vessel.get('MV'))
        if not key:
            continue

        date = normalise_date(vessel.get('Vessel Open Date'))

        if key not in existing_vessels:
            existing_vessels[key] = date
            vessels.append(vessel)
        else:
            if existing_vessels[key] != date:
                existing_vessels[key] = date
                vessels.append(vessel)

    if vessels:
        save_duplicates(existing_vessels)

    return vessels


def delete_duplicates():
    global existing_vessels
    existing_vessels = {}
    with open(duplicates_file, "w", encoding="utf-8") as f:
        json.dump({}, f)  # empty dict


def process_email(email_address,folder,excel_path,csv_dict,worker):

    global email_ids
    start_time = datetime.now()
    pythoncom.CoInitialize()
    try:
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

        for account in outlook.Folders:
            if account.Name.upper() == email_address.upper():
                for subfolder in account.Folders:
                    if subfolder.Name.upper() == folder.upper():
                        folder = subfolder
                        break
                else:
                    return
                break
        else:
            return

        store_id = folder.StoreID
        api_failures = 0  # consecutive API failures (circuit breaker for a real outage)

        while True:
            if not worker.running:
                return

            refresh_access_state()

            # Phase 1 — snapshot new (unseen) EntryIDs quickly, without holding the live
            # enumerator across the slow per-email work below.
            entry_ids = []
            try:
                messages = folder.Items
                messages.Sort("[ReceivedTime]", True)
                message = messages.GetFirst()
                while message:
                    if not worker.running:
                        return
                    if hasattr(message, 'ReceivedTime'):
                        if message.ReceivedTime.replace(tzinfo=None) < start_time:
                            break
                        try:
                            eid = message.EntryID
                        except Exception:
                            message = messages.GetNext()
                            continue
                        with _email_ids_lock:
                            seen = eid in email_ids
                        if not seen:
                            entry_ids.append(eid)
                    message = messages.GetNext()
            except Exception as e:
                logger.warning(f"Listen snapshot interrupted ({e})")

            # Phase 2 — process each from a fresh item handle; one failure skips, not aborts.
            for entry_id in reversed(entry_ids):  # oldest-first, matching arrival order
                if not worker.running:
                    return
                try:
                    message = outlook.GetItemFromID(entry_id, store_id)
                    received_time = message.ReceivedTime
                    email_body = message.Body
                    email_subject = message.Subject
                    sender_email = message.SenderEmailAddress
                except Exception as e:
                    logger.warning(f"Skipping email (fetch failed): {e}")
                    continue  # don't mark seen — retry next cycle

                with _email_ids_lock:
                    email_ids.add(entry_id)
                save_email_ids()

                limit_hit = False
                rows = []
                excel_vessels = None
                try:
                    if is_relevant_email(email_subject, email_body):
                        logger.info(f"Processing email from: {sender_email} with subject: {email_subject}")
                        preprocessed_body = get_first_n_lines(email_body)
                        if not preprocessed_body:
                            append_error_message(excel_path, sender_email, email_subject)
                        else:
                            if not access_allowed():
                                limit_hit = True
                            else:
                                extracted_details = extract_details_from_email(preprocessed_body, csv_dict)
                                api_failures = 0  # a call succeeded → reset breaker
                                valid_vessels = filter_data(extracted_details)
                                vessels = detect_duplicates(valid_vessels)
                                if vessels:
                                    for vessel in vessels:
                                        vessel['Sender'] = sender_email
                                        vessel['Subject'] = email_subject
                                        vessel['Received Time'] = format_received_time(received_time)
                                    ves = len(vessels)
                                    rows = [{
                                        "sender": sender_email,
                                        "subject": email_subject,
                                        "received_time": format_received_time(received_time),
                                        "ves": ves,
                                        "vessel_data": vessel,
                                    } for vessel in vessels]
                                    excel_vessels = vessels
                except APIError as e:
                    api_failures += 1
                    logger.warning(f"API failed for this email; skipping (consecutive={api_failures}): {e}")
                    if api_failures >= 5:
                        yield {"type": "api_error", "error_key": "proxy_error_generic"}
                        return
                    continue
                except Exception as e:
                    logger.warning(f"Skipping email (processing failed): {e}")
                    continue

                if limit_hit:
                    yield {"type": "limit_reached"}
                    return
                for row in rows:
                    yield row
                if excel_vessels:
                    if is_excel_open(excel_path):
                        yield {"type": "excel_locked"}
                        while is_excel_open(excel_path):
                            if not worker.running:
                                return
                            time.sleep(2)
                        yield {"type": "excel_unlocked"}
                    append_data_excel(excel_path, excel_vessels, None, False)

            for _ in range(10):
                if not worker.running:
                    return
                time.sleep(1)
    finally:
        pythoncom.CoUninitialize()




def night_extraction(specific_datetime, email_address, folder, excel_path, csv_dict, worker):

    pythoncom.CoInitialize()
    try:
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

        for account in outlook.Folders:
            if account.Name.upper() == email_address.upper():
                for subfolder in account.Folders:
                    if subfolder.Name.upper() == folder.upper():
                        folder = subfolder
                        break
                else:
                    return
                break
        else:
            return

        store_id = folder.StoreID
        refresh_access_state()

        # Phase 1 — snapshot relevant EntryIDs quickly (no API calls; minimal time holding
        # the live COM enumerator, so incoming mail / a dropped connection can't break us
        # mid-run the way iterating across slow OpenAI calls did).
        entry_ids = []
        try:
            messages = folder.Items
            messages.Sort("[ReceivedTime]", True)
            message = messages.GetFirst()
            while message:
                if not worker.running:
                    return
                if hasattr(message, 'ReceivedTime'):
                    received_time = message.ReceivedTime
                    if received_time.replace(tzinfo=None) <= specific_datetime.replace(tzinfo=None):
                        break
                    try:
                        entry_ids.append(message.EntryID)
                    except Exception:
                        pass
                message = messages.GetNext()
        except Exception as e:
            logger.warning(f"Snapshot interrupted ({e}); proceeding with {len(entry_ids)} emails collected")

        # Phase 2 — process each from a fresh item handle. A failure on one email skips it
        # rather than aborting the whole run.
        processed_emails = []
        ves = 0
        api_failures = 0  # consecutive API failures (circuit breaker for a real outage)
        for entry_id in entry_ids:
            if not worker.running:
                return

            limit_hit = False
            rows = []
            try:
                message = outlook.GetItemFromID(entry_id, store_id)
                received_time = message.ReceivedTime
                email_body = message.Body
                email_subject = message.Subject
                sender_email = message.SenderEmailAddress

                if not is_relevant_email(email_subject, email_body):
                    continue

                logger.info(f"Processing email from: {sender_email} with subject: {email_subject}")
                preprocessed_body = get_first_n_lines(email_body)
                if not preprocessed_body:
                    append_error_message(excel_path, sender_email, email_subject)
                    continue

                if not access_allowed():
                    limit_hit = True
                else:
                    extracted_details = extract_details_from_email(preprocessed_body, csv_dict)
                    api_failures = 0  # a call succeeded → reset breaker
                    valid_vessels = filter_data(extracted_details)
                    vessels = detect_duplicates(valid_vessels)
                    if vessels:
                        for vessel in vessels:
                            vessel['Sender'] = sender_email
                            vessel['Subject'] = email_subject
                            vessel['Received Time'] = format_received_time(received_time)
                        processed_emails.extend(vessels)
                        ves += len(vessels)
                        rows = [{
                            "sender": sender_email,
                            "subject": email_subject,
                            "received_time": format_received_time(received_time),
                            "ves": ves,
                            "vessel_data": vessel,
                        } for vessel in vessels]
            except APIError as e:
                api_failures += 1
                logger.warning(f"API failed for this email; skipping (consecutive={api_failures}): {e}")
                if api_failures >= 5:
                    yield {"type": "api_error", "error_key": "proxy_error_generic"}
                    return
                continue
            except Exception as e:
                logger.warning(f"Skipping email (processing failed): {e}")
                continue

            if limit_hit:
                yield {"type": "limit_reached"}
                return
            for row in rows:
                yield row

        if processed_emails:
            if is_excel_open(excel_path):
                logger.warning("Output Excel file is locked (open in Excel / syncing) — waiting to write results")
                yield {"type": "excel_locked"}
                waited = 0
                while is_excel_open(excel_path):
                    if not worker.running:
                        return
                    time.sleep(2)
                    waited += 2
                    if waited >= 300:  # give up after ~5 min rather than hang forever
                        logger.error("Excel file still locked after 5 min — aborting write")
                        yield {"type": "api_error", "error_key": "proxy_error_generic"}
                        return
                yield {"type": "excel_unlocked"}
            logger.info(f"Writing {len(processed_emails)} vessels to Excel")
            append_data_excel(excel_path, processed_emails, specific_datetime, True)
            logger.info("Extraction complete — results written to Excel")
            return True
        else:
            return False
    finally:
        pythoncom.CoUninitialize()
